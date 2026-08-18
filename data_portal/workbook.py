from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill(
    "solid",
    fgColor="173D7A",
)
TITLE_FILL = PatternFill(
    "solid",
    fgColor="0F2D5C",
)
ACCENT_FILL = PatternFill(
    "solid",
    fgColor="F2A900",
)
REQUIRED_FILL = PatternFill(
    "solid",
    fgColor="FFF2CC",
)
WHITE_FONT = Font(
    color="FFFFFF",
    bold=True,
)


def style_header(sheet, row=1):
    for cell in sheet[row]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    sheet.freeze_panes = f"A{row + 1}"
    sheet.auto_filter.ref = sheet.dimensions


def set_widths(sheet, widths):
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def build_student_import_workbook(portal):
    workbook = Workbook()

    instructions = workbook.active
    instructions.title = "INSTRUCOES"

    schools = workbook.create_sheet("ESCOLAS")
    classrooms = workbook.create_sheet("TURMAS")
    students = workbook.create_sheet("ALUNOS")
    lists = workbook.create_sheet("LISTAS")

    application = portal.application
    assessment = application.assessment

    instructions.merge_cells("A1:F1")
    instructions["A1"] = "DINÂMICA SIMULADOS — PREPARAÇÃO DOS DADOS"
    instructions["A1"].fill = TITLE_FILL
    instructions["A1"].font = Font(
        color="FFFFFF",
        bold=True,
        size=16,
    )
    instructions["A1"].alignment = Alignment(
        horizontal="center",
    )

    instructions["A3"] = "Aplicação"
    instructions["B3"] = application.title
    instructions["A4"] = "Código"
    instructions["B4"] = application.code
    instructions["A5"] = "Município"
    instructions["B5"] = str(application.municipality)
    instructions["A6"] = "Prova"
    instructions["B6"] = assessment.title
    instructions["A7"] = "Data"
    instructions["B7"] = application.application_date
    instructions["B7"].number_format = "yyyy-mm-dd"

    instructions["A9"] = "ORIENTAÇÕES"
    instructions["A9"].fill = ACCENT_FILL
    instructions["A9"].font = Font(bold=True)

    guidance = [
        "Não altere os nomes das abas.",
        "Não altere os títulos das colunas.",
        "Preencha primeiro ESCOLAS, depois TURMAS e ALUNOS.",
        "Use os mesmos códigos de escola e turma em todas as abas.",
        "Campos marcados com * são obrigatórios.",
        "Não inclua CPF, endereço ou telefone pessoal do aluno.",
        "A matrícula precisa ser única dentro de cada escola.",
        "Datas devem estar no formato DD/MM/AAAA.",
        "Disciplinas devem ser separadas por ponto e vírgula.",
        "Exemplo de disciplinas: PORT;MAT.",
    ]

    for row, guidance_text in enumerate(guidance, start=10):
        instructions.cell(
            row=row,
            column=1,
            value=f"{row - 9}.",
        )
        instructions.cell(
            row=row,
            column=2,
            value=guidance_text,
        )

    instructions.column_dimensions["A"].width = 18
    instructions.column_dimensions["B"].width = 82

    school_headers = [
        "codigo_escola*",
        "codigo_inep",
        "nome_escola*",
        "endereco",
    ]
    schools.append(school_headers)
    schools.append([
        "ESC001",
        "23012345",
        "Escola Municipal Horizonte",
        "Rua Exemplo, 100",
    ])
    style_header(schools)
    set_widths(
        schools,
        {
            "A": 20,
            "B": 18,
            "C": 42,
            "D": 45,
        },
    )

    classroom_headers = [
        "codigo_turma*",
        "codigo_escola*",
        "ano_letivo*",
        "codigo_serie*",
        "nome_turma*",
        "turno*",
        "sala",
        "disciplinas*",
    ]
    classrooms.append(classroom_headers)
    classrooms.append([
        "ESC001-EF09-A",
        "ESC001",
        assessment.academic_year.year,
        "EF09",
        "A",
        "Manhã",
        "Sala 01",
        "PORT;MAT",
    ])
    style_header(classrooms)
    set_widths(
        classrooms,
        {
            "A": 24,
            "B": 20,
            "C": 16,
            "D": 18,
            "E": 18,
            "F": 18,
            "G": 18,
            "H": 24,
        },
    )

    student_headers = [
        "codigo_escola*",
        "codigo_turma*",
        "matricula*",
        "nome_completo*",
        "data_nascimento",
        "necessidade_atendimento",
        "observacao_aplicacao",
    ]
    students.append(student_headers)
    students.append([
        "ESC001",
        "ESC001-EF09-A",
        "20260001",
        "Ana Beatriz da Silva",
        "",
        "",
        "",
    ])
    style_header(students)
    set_widths(
        students,
        {
            "A": 20,
            "B": 24,
            "C": 22,
            "D": 42,
            "E": 20,
            "F": 30,
            "G": 38,
        },
    )

    for cell in students["C"]:
        cell.number_format = "@"

    for cell in students["E"]:
        if cell.row > 1:
            cell.number_format = "dd/mm/yyyy"

    lists.append([
        "TURNOS",
        "SERIES",
        "DISCIPLINAS",
    ])

    shifts = [
        "Manhã",
        "Tarde",
        "Noite",
        "Integral",
    ]

    allowed_grades = list(
        assessment.grades.filter(
            is_active=True,
        ).values_list(
            "code",
            flat=True,
        )
    )

    if assessment.components.exists():
        subject_codes = list(
            assessment.components.filter(
                is_active=True,
            ).values_list(
                "subject__code",
                flat=True,
            )
        )
    elif assessment.subject_id:
        subject_codes = [assessment.subject.code]
    else:
        subject_codes = []

    maximum_rows = max(
        len(shifts),
        len(allowed_grades),
        len(subject_codes),
    )

    for index in range(maximum_rows):
        lists.append([
            shifts[index] if index < len(shifts) else "",
            (
                allowed_grades[index]
                if index < len(allowed_grades)
                else ""
            ),
            (
                subject_codes[index]
                if index < len(subject_codes)
                else ""
            ),
        ])

    style_header(lists)
    set_widths(
        lists,
        {
            "A": 20,
            "B": 20,
            "C": 24,
        },
    )

    shift_validation = DataValidation(
        type="list",
        formula1=(
            f"'LISTAS'!$A$2:$A${len(shifts) + 1}"
        ),
        allow_blank=False,
    )
    classrooms.add_data_validation(shift_validation)
    shift_validation.add("F2:F1000")

    if allowed_grades:
        grade_validation = DataValidation(
            type="list",
            formula1=(
                "'LISTAS'!$B$2:"
                f"$B${len(allowed_grades) + 1}"
            ),
            allow_blank=False,
        )
        classrooms.add_data_validation(grade_validation)
        grade_validation.add("D2:D1000")

    for sheet in (
        schools,
        classrooms,
        students,
    ):
        sheet.sheet_view.showGridLines = False

        for cell in sheet[1]:
            if cell.value and str(cell.value).endswith("*"):
                cell.fill = REQUIRED_FILL
                cell.font = Font(
                    color="1F2937",
                    bold=True,
                )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output