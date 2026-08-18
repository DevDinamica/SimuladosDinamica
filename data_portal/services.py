from datetime import datetime
from pathlib import Path

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from academics.models import (
    Classroom,
    Enrollment,
    Grade,
    Student,
    Subject,
)
from applications.models import ApplicationClassroom
from institutions.models import School

from .models import DataPreparationPortal, SpreadsheetUpload


REQUIRED_SHEETS = {
    "ESCOLAS",
    "TURMAS",
    "ALUNOS",
}

REQUIRED_HEADERS = {
    "ESCOLAS": {
        "codigo_escola",
        "nome_escola",
    },
    "TURMAS": {
        "codigo_turma",
        "codigo_escola",
        "ano_letivo",
        "codigo_serie",
        "nome_turma",
        "turno",
        "disciplinas",
    },
    "ALUNOS": {
        "codigo_escola",
        "codigo_turma",
        "matricula",
        "nome_completo",
    },
}

SHIFT_MAP = {
    "MANHÃ": Classroom.Shift.MORNING,
    "MANHA": Classroom.Shift.MORNING,
    "MATUTINO": Classroom.Shift.MORNING,
    "TARDE": Classroom.Shift.AFTERNOON,
    "VESPERTINO": Classroom.Shift.AFTERNOON,
    "NOITE": Classroom.Shift.EVENING,
    "NOTURNO": Classroom.Shift.EVENING,
    "INTEGRAL": Classroom.Shift.FULL_TIME,
}


def normalize_header(value):
    if value is None:
        return ""

    return (
        str(value)
        .strip()
        .lower()
        .replace("*", "")
        .replace(" ", "_")
    )


def normalize_text(value):
    if value is None:
        return ""

    return str(value).strip()


def read_sheet_rows(sheet):
    headers = [
        normalize_header(cell.value)
        for cell in sheet[1]
    ]

    rows = []

    for row_number, values in enumerate(
        sheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):
        if not any(
            value not in (None, "")
            for value in values
        ):
            continue

        row = {
            headers[index]: values[index]
            if index < len(values)
            else None
            for index in range(len(headers))
            if headers[index]
        }
        row["_row"] = row_number
        rows.append(row)

    return headers, rows


def validate_upload(upload):
    report = {
        "errors": [],
        "warnings": [],
    }

    try:
        workbook = load_workbook(
            upload.file.path,
            read_only=True,
            data_only=True,
        )
    except Exception:
        report["errors"].append({
            "sheet": "ARQUIVO",
            "row": None,
            "message": (
                "Não foi possível abrir a planilha. "
                "Baixe novamente o modelo oficial."
            ),
        })
        return finalize_validation(upload, report, {}, {}, {})

    missing_sheets = REQUIRED_SHEETS - set(
        workbook.sheetnames
    )

    for sheet_name in sorted(missing_sheets):
        report["errors"].append({
            "sheet": sheet_name,
            "row": None,
            "message": "Aba obrigatória não encontrada.",
        })

    if missing_sheets:
        return finalize_validation(upload, report, {}, {}, {})

    parsed = {}

    for sheet_name in REQUIRED_SHEETS:
        headers, rows = read_sheet_rows(
            workbook[sheet_name]
        )
        parsed[sheet_name] = rows

        missing_headers = (
            REQUIRED_HEADERS[sheet_name]
            - set(headers)
        )

        for header in sorted(missing_headers):
            report["errors"].append({
                "sheet": sheet_name,
                "row": 1,
                "message": (
                    f"Coluna obrigatória ausente: {header}."
                ),
            })

    schools = parsed["ESCOLAS"]
    classrooms = parsed["TURMAS"]
    students = parsed["ALUNOS"]

    school_codes = {}
    classroom_codes = {}
    student_keys = set()

    for row in schools:
        row_number = row["_row"]
        code = normalize_text(
            row.get("codigo_escola")
        )
        name = normalize_text(
            row.get("nome_escola")
        )

        if not code:
            add_error(
                report,
                "ESCOLAS",
                row_number,
                "Código da escola é obrigatório.",
            )
            continue

        if not name:
            add_error(
                report,
                "ESCOLAS",
                row_number,
                "Nome da escola é obrigatório.",
            )

        if code in school_codes:
            add_error(
                report,
                "ESCOLAS",
                row_number,
                f"Código de escola duplicado: {code}.",
            )
        else:
            school_codes[code] = row

    allowed_grade_codes = set(
        upload.portal.application.assessment.grades.values_list(
            "code",
            flat=True,
        )
    )

    allowed_subject_codes = set(
        upload.portal.application.assessment.components.filter(
            is_active=True,
        ).values_list(
            "subject__code",
            flat=True,
        )
    )

    if not allowed_subject_codes:
        subject = upload.portal.application.assessment.subject

        if subject:
            allowed_subject_codes = {subject.code}

    expected_year = (
        upload.portal.application.assessment.academic_year.year
    )

    for row in classrooms:
        row_number = row["_row"]
        code = normalize_text(
            row.get("codigo_turma")
        )
        school_code = normalize_text(
            row.get("codigo_escola")
        )
        grade_code = normalize_text(
            row.get("codigo_serie")
        ).upper()
        shift_name = normalize_text(
            row.get("turno")
        ).upper()

        try:
            academic_year = int(
                row.get("ano_letivo")
            )
        except (TypeError, ValueError):
            academic_year = None

        if not code:
            add_error(
                report,
                "TURMAS",
                row_number,
                "Código da turma é obrigatório.",
            )
            continue

        if code in classroom_codes:
            add_error(
                report,
                "TURMAS",
                row_number,
                f"Código de turma duplicado: {code}.",
            )
        else:
            classroom_codes[code] = row

        if school_code not in school_codes:
            add_error(
                report,
                "TURMAS",
                row_number,
                (
                    f"Escola {school_code} não encontrada "
                    "na aba ESCOLAS."
                ),
            )

        if grade_code not in allowed_grade_codes:
            add_error(
                report,
                "TURMAS",
                row_number,
                (
                    f"Série {grade_code} não está autorizada "
                    "para este simulado."
                ),
            )

        if academic_year != expected_year:
            add_error(
                report,
                "TURMAS",
                row_number,
                (
                    f"O ano letivo deve ser {expected_year}."
                ),
            )

        if shift_name not in SHIFT_MAP:
            add_error(
                report,
                "TURMAS",
                row_number,
                f"Turno não reconhecido: {shift_name}.",
            )

        subject_codes = {
            item.strip().upper()
            for item in normalize_text(
                row.get("disciplinas")
            ).split(";")
            if item.strip()
        }

        invalid_subjects = (
            subject_codes - allowed_subject_codes
        )

        if invalid_subjects:
            add_error(
                report,
                "TURMAS",
                row_number,
                (
                    "Disciplinas não autorizadas: "
                    f"{', '.join(sorted(invalid_subjects))}."
                ),
            )

    for row in students:
        row_number = row["_row"]
        school_code = normalize_text(
            row.get("codigo_escola")
        )
        classroom_code = normalize_text(
            row.get("codigo_turma")
        )
        registration = normalize_text(
            row.get("matricula")
        )
        full_name = normalize_text(
            row.get("nome_completo")
        )

        if school_code not in school_codes:
            add_error(
                report,
                "ALUNOS",
                row_number,
                (
                    f"Escola {school_code} não encontrada."
                ),
            )

        if classroom_code not in classroom_codes:
            add_error(
                report,
                "ALUNOS",
                row_number,
                (
                    f"Turma {classroom_code} não encontrada."
                ),
            )
        elif (
            normalize_text(
                classroom_codes[classroom_code].get(
                    "codigo_escola"
                )
            )
            != school_code
        ):
            add_error(
                report,
                "ALUNOS",
                row_number,
                "A turma não pertence à escola informada.",
            )

        if not registration:
            add_error(
                report,
                "ALUNOS",
                row_number,
                "Matrícula é obrigatória.",
            )

        if not full_name:
            add_error(
                report,
                "ALUNOS",
                row_number,
                "Nome completo é obrigatório.",
            )

        student_key = (
            school_code,
            registration,
        )

        if student_key in student_keys:
            add_error(
                report,
                "ALUNOS",
                row_number,
                (
                    f"Matrícula duplicada na escola: "
                    f"{registration}."
                ),
            )
        else:
            student_keys.add(student_key)

    return finalize_validation(
        upload,
        report,
        schools,
        classrooms,
        students,
    )


def add_error(report, sheet, row, message):
    report["errors"].append({
        "sheet": sheet,
        "row": row,
        "message": message,
    })


def finalize_validation(
    upload,
    report,
    schools,
    classrooms,
    students,
):
    upload.school_count = len(schools)
    upload.classroom_count = len(classrooms)
    upload.student_count = len(students)
    upload.error_count = len(
        report["errors"]
    )
    upload.warning_count = len(
        report["warnings"]
    )
    upload.validation_report = report

    if report["errors"]:
        upload.status = SpreadsheetUpload.Status.INVALID
        upload.portal.status = (
            DataPreparationPortal.Status.HAS_ERRORS
        )
    else:
        upload.status = SpreadsheetUpload.Status.VALID
        upload.portal.status = (
            DataPreparationPortal.Status.READY
        )

    upload.save()
    upload.portal.save(
        update_fields=[
            "status",
            "updated_at",
        ]
    )

    return report


@transaction.atomic
def import_validated_upload(upload):
    if upload.status != SpreadsheetUpload.Status.VALID:
        raise ValidationError(
            "Somente planilhas validadas podem ser importadas."
        )

    workbook = load_workbook(
        upload.file.path,
        read_only=True,
        data_only=True,
    )

    _, school_rows = read_sheet_rows(
        workbook["ESCOLAS"]
    )
    _, classroom_rows = read_sheet_rows(
        workbook["TURMAS"]
    )
    _, student_rows = read_sheet_rows(
        workbook["ALUNOS"]
    )

    application = upload.portal.application
    municipality = application.municipality
    academic_year = application.assessment.academic_year

    schools_by_code = {}
    classrooms_by_code = {}

    created_schools = 0
    created_classrooms = 0
    created_students = 0
    created_enrollments = 0

    for row in school_rows:
        school_code = normalize_text(
            row.get("codigo_escola")
        )
        inep_code = normalize_text(
            row.get("codigo_inep")
        )
        name = normalize_text(
            row.get("nome_escola")
        )
        address = normalize_text(
            row.get("endereco")
        )

        school = None

        if inep_code:
            school = School.objects.filter(
                municipality=municipality,
                inep_code=inep_code,
            ).first()

        if school is None:
            school, created = School.objects.get_or_create(
                municipality=municipality,
                name=name,
                defaults={
                    "inep_code": inep_code,
                    "address": address,
                    "is_active": True,
                },
            )

            if created:
                created_schools += 1
        else:
            school.name = name
            school.address = address
            school.is_active = True
            school.save()

        schools_by_code[school_code] = school

    for row in classroom_rows:
        classroom_code = normalize_text(
            row.get("codigo_turma")
        )
        school_code = normalize_text(
            row.get("codigo_escola")
        )
        grade_code = normalize_text(
            row.get("codigo_serie")
        ).upper()
        name = normalize_text(
            row.get("nome_turma")
        )
        shift_name = normalize_text(
            row.get("turno")
        ).upper()

        school = schools_by_code[school_code]
        grade = Grade.objects.get(
            code=grade_code,
        )

        classroom, created = Classroom.objects.get_or_create(
            school=school,
            academic_year=academic_year,
            grade=grade,
            name=name,
            shift=SHIFT_MAP[shift_name],
            defaults={
                "is_active": True,
            },
        )

        if created:
            created_classrooms += 1

        subject_codes = {
            item.strip().upper()
            for item in normalize_text(
                row.get("disciplinas")
            ).split(";")
            if item.strip()
        }

        subjects = Subject.objects.filter(
            code__in=subject_codes,
        )
        classroom.subjects.add(*subjects)

        classrooms_by_code[classroom_code] = classroom

        ApplicationClassroom.objects.get_or_create(
            application=application,
            classroom=classroom,
            defaults={
                "room_name": normalize_text(
                    row.get("sala")
                ),
                "is_active": True,
            },
        )

    for row in student_rows:
        school_code = normalize_text(
            row.get("codigo_escola")
        )
        classroom_code = normalize_text(
            row.get("codigo_turma")
        )
        registration = normalize_text(
            row.get("matricula")
        )
        full_name = normalize_text(
            row.get("nome_completo")
        )

        birth_date = row.get(
            "data_nascimento"
        )

        if isinstance(birth_date, str):
            try:
                birth_date = datetime.strptime(
                    birth_date,
                    "%d/%m/%Y",
                ).date()
            except ValueError:
                birth_date = None

        school = schools_by_code[school_code]
        classroom = classrooms_by_code[
            classroom_code
        ]

        student, created = Student.objects.update_or_create(
            school=school,
            registration_code=registration,
            defaults={
                "full_name": full_name,
                "birth_date": birth_date,
                "is_active": True,
            },
        )

        if created:
            created_students += 1

        _, enrollment_created = (
            Enrollment.objects.get_or_create(
                student=student,
                classroom=classroom,
                defaults={
                    "status": Enrollment.Status.ACTIVE,
                },
            )
        )

        if enrollment_created:
            created_enrollments += 1

    upload.status = SpreadsheetUpload.Status.IMPORTED
    upload.imported_at = timezone.now()
    upload.save()

    portal = upload.portal
    portal.status = DataPreparationPortal.Status.IMPORTED
    portal.imported_at = timezone.now()
    portal.is_active = False
    portal.save()

    application.status = (
        application.Status.PREPARING
    )
    application.save()

    return {
        "schools_created": created_schools,
        "classrooms_created": created_classrooms,
        "students_created": created_students,
        "enrollments_created": created_enrollments,
    }